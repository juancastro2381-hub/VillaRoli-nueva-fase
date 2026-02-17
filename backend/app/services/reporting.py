import io
from typing import List, Dict, Any
from datetime import date
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.db.models import Booking


class ReportingService:
    @staticmethod
    def generate_bookings_pdf(bookings: List[Booking]) -> bytes:
        """Legacy booking report - no financial data"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Reporte de Reservas - Villa Roli", styles['Title']))
        elements.append(Paragraph(f"Generado: {date.today()}", styles['Normal']))
        elements.append(Paragraph(" ", styles['Normal']))
        
        # Updated Headers
        data = [['ID', 'Entrada', 'Salida', 'Plan', 'Pax', 'Estado', 'Canal', 'Excepción']]
        
        # Rows
        for b in bookings:
            payment_status = b.payments[0].status if b.payments else "N/A"
            channel = "Admin" if b.created_by_admin_id else "Online"
            override = "SÍ" if b.is_override else "NO"
            
            data.append([
                str(b.id),
                str(b.check_in),
                str(b.check_out),
                b.policy_type.value[:10], # Truncate for PDF
                str(b.guest_count),
                b.status.value,
                channel,
                override
            ])
            
        # Table Style
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_bookings_xlsx(bookings: List[Booking]) -> bytes:
        """Legacy booking report - no financial data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reservas"
        
        # Headers
        headers = ['ID', 'Check-In', 'Check-Out', 'Plan', 'Huéspedes', 'Estado', 'Pago', 
                   'Método Pago', 'Canal', 'Excepción', 'Motivo Excepción', 'Admin ID']
        ws.append(headers)
        
        for b in bookings:
            payment = b.payments[0] if b.payments else None
            payment_status = payment.status if payment else "N/A"
            payment_method = payment.payment_method if payment else "N/A"
            channel = "Manual Admin" if b.created_by_admin_id else "Online"
            
            ws.append([
                b.id,
                b.check_in,
                b.check_out,
                b.policy_type.value,
                b.guest_count,
                b.status.value,
                payment_status,
                payment_method,
                channel,
                "SÍ" if b.is_override else "NO",
                b.override_reason or "",
                b.created_by_admin_id or ""
            ])
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_financial_report_pdf(details: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
        """
        Generate comprehensive financial report PDF.
        Uses FinancialService data for consistency.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph("Reporte Financiero - Villa Roli", styles['Title']))
        elements.append(Paragraph(f"Generado: {date.today()}", styles['Normal']))
        
        # Date Range
        if summary['date_range']['from'] and summary['date_range']['to']:
            elements.append(Paragraph(
                f"Período: {summary['date_range']['from']} a {summary['date_range']['to']}", 
                styles['Normal']
            ))
        elements.append(Spacer(1, 0.2*inch))
        
        # Summary Section
        elements.append(Paragraph("<b>Resumen Financiero</b>", styles['Heading2']))
        summary_text = f"""
        <b>Ingreso Total:</b> ${summary['total_revenue']:,.2f} COP<br/>
        <b>Reservas Confirmadas:</b> {summary['total_bookings_confirmed']}<br/>
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Detailed Table
        elements.append(Paragraph("<b>Detalle de Reservas</b>", styles['Heading2']))
        
        data = [['ID', 'Cliente', 'Check-In', 'Check-Out', 'Plan', 'Pax', 
                 'Monto', 'Método', 'Canal', 'Confirmado']]
        
        for detail in details:
            data.append([
                str(detail['booking_id']),
                detail['guest_name'] or 'N/A',
                detail['check_in'],
                detail['check_out'],
                detail['plan'][:15],  # Truncate
                str(detail['guest_count']),
                f"${detail['amount']:,.2f}",
                detail['payment_method'][:10],
                detail['channel'],
                detail['confirmed_at'] or 'N/A'
            ])
        
        # Table Style
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),  # Amount column
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Revenue Breakdown
        elements.append(Paragraph("<b>Desglose por Canal</b>", styles['Heading3']))
        for channel, amount in summary['revenue_by_channel'].items():
            elements.append(Paragraph(f"{channel.capitalize()}: ${amount:,.2f} COP", styles['Normal']))
        
        doc.build(elements)
        
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_financial_report_xlsx(details: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
        """
        Generate comprehensive financial report XLSX.
        Uses FinancialService data for consistency.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Financiero"
        
        # Title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "Reporte Financiero - Villa Roli"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Date info
        ws.merge_cells('A2:J2')
        date_cell = ws['A2']
        date_cell.value = f"Generado: {date.today()}"
        date_cell.alignment = Alignment(horizontal='center')
        
        # Date range
        if summary['date_range']['from'] and summary['date_range']['to']:
            ws.merge_cells('A3:J3')
            range_cell = ws['A3']
            range_cell.value = f"Período: {summary['date_range']['from']} a {summary['date_range']['to']}"
            range_cell.alignment = Alignment(horizontal='center')
            start_row = 5
        else:
            start_row = 4
        
        # Summary Section
        ws[f'A{start_row}'] = "RESUMEN FINANCIERO"
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        
        ws[f'A{start_row+1}'] = "Ingreso Total:"
        ws[f'B{start_row+1}'] = f"${summary['total_revenue']:,.2f} COP"
        ws[f'B{start_row+1}'].font = Font(bold=True, size=14, color="006100")
        
        ws[f'A{start_row+2}'] = "Reservas Confirmadas:"
        ws[f'B{start_row+2}'] = summary['total_bookings_confirmed']
        
        # Headers for detail table
        header_row = start_row + 4
        headers = ['ID', 'Cliente', 'Email', 'Check-In', 'Check-Out', 'Plan', 'Pax', 
                   'Monto', 'Método Pago', 'Canal', 'Confirmado']
        ws.append([''])  # Spacer
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row+1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for detail in details:
            ws.append([
                detail['booking_id'],
                detail['guest_name'] or 'N/A',
                detail['guest_email'] or 'N/A',
                detail['check_in'],
                detail['check_out'],
                detail['plan'],
                detail['guest_count'],
                detail['amount'],
                detail['payment_method'],
                detail['channel'],
                detail['confirmed_at'] or 'N/A'
            ])
        
        # Format amount column (column H)
        for row in range(header_row+2, ws.max_row+1):
            amount_cell = ws.cell(row=row, column=8)
            amount_cell.number_format = '$#,##0.00'
        
        # Summary row
        summary_row = ws.max_row + 2
        ws[f'G{summary_row}'] = "TOTAL:"
        ws[f'G{summary_row}'].font = Font(bold=True)
        ws[f'H{summary_row}'] = summary['total_revenue']
        ws[f'H{summary_row}'].font = Font(bold=True, size=12)
        ws[f'H{summary_row}'].number_format = '$#,##0.00'
        ws[f'H{summary_row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
